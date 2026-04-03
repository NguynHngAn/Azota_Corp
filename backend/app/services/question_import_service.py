import csv
from io import BytesIO, StringIO

from docx import Document
from openpyxl import load_workbook

from app.models.exam import QuestionType
from app.models.question_bank import QuestionDifficulty
from app.schemas.question_bank_schema import BankAnswerOptionCreate, BankQuestionCreate


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "correct", "x"}


def _parse_question_type(value: object) -> QuestionType:
    text = str(value or "").strip().lower()
    if text in {"multiple", "multiple_choice", "multiple-choice", "mcq-multi"}:
        return QuestionType.multiple_choice
    return QuestionType.single_choice


def _parse_difficulty(value: object) -> QuestionDifficulty:
    text = str(value or "").strip().lower()
    if text == "easy":
        return QuestionDifficulty.easy
    if text == "hard":
        return QuestionDifficulty.hard
    return QuestionDifficulty.medium


def _parse_tabular_rows(rows: list[list[object]]) -> list[BankQuestionCreate]:
    if not rows:
        return []
    # Excel can include a UTF-8 BOM on the first header cell (e.g. "\ufeffquestion"),
    # which would break our exact header matching.
    headers = [
        str(cell or "")
        .replace("\ufeff", "")
        .strip()
        .lower()
        for cell in rows[0]
    ]
    required = {"question", "type", "option_1", "option_2", "correct_options"}
    if not required.issubset(set(headers)):
        missing = ", ".join(sorted(required - set(headers)))
        raise ValueError(f"Missing required columns: {missing}")

    header_index = {name: idx for idx, name in enumerate(headers)}
    parsed: list[BankQuestionCreate] = []
    for row_number, row in enumerate(rows[1:], start=2):
        padded_row = list(row) + [None] * max(0, len(headers) - len(row))
        question_text = str(padded_row[header_index["question"]] or "").strip()
        if not question_text:
            continue
        question_type = _parse_question_type(padded_row[header_index["type"]])
        difficulty = _parse_difficulty(padded_row[header_index["difficulty"]]) if "difficulty" in header_index else QuestionDifficulty.medium
        explanation = str(padded_row[header_index["explanation"]] or "").strip() if "explanation" in header_index else ""
        tags = []
        if "tags" in header_index:
            tags = [part.strip() for part in str(padded_row[header_index["tags"]] or "").split(",") if part.strip()]

        option_texts: list[str] = []
        for key in ("option_1", "option_2", "option_3", "option_4", "option_5", "option_6"):
            if key in header_index:
                text = str(padded_row[header_index[key]] or "").strip()
                if text:
                    option_texts.append(text)
        if len(option_texts) < 2:
            raise ValueError(f"Row {row_number}: at least 2 options are required")

        correct_raw = str(padded_row[header_index["correct_options"]] or "").strip()
        if not correct_raw:
            raise ValueError(f"Row {row_number}: correct_options is required")
        correct_indexes = {
            int(part.strip()) - 1
            for part in correct_raw.split(",")
            if part.strip().isdigit()
        }
        if not correct_indexes:
            raise ValueError(f"Row {row_number}: correct_options must contain option indexes like 1 or 1,3")

        options = [
            BankAnswerOptionCreate(
                order_index=index,
                text=text,
                is_correct=index in correct_indexes,
            )
            for index, text in enumerate(option_texts)
        ]
        parsed.append(
            BankQuestionCreate(
                question_type=question_type,
                text=question_text,
                explanation=explanation or None,
                difficulty=difficulty,
                is_active=True,
                options=options,
                tags=tags,
            )
        )
    return parsed


def parse_excel_questions(content: bytes) -> list[BankQuestionCreate]:
    workbook = load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook.active
    rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return _parse_tabular_rows(rows)


def parse_csv_questions(content: bytes) -> list[BankQuestionCreate]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(StringIO(text))
    rows = [list(row) for row in reader]
    return _parse_tabular_rows(rows)


def parse_word_questions(content: bytes) -> list[BankQuestionCreate]:
    document = Document(BytesIO(content))
    lines = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    questions: list[BankQuestionCreate] = []
    current: dict | None = None
    for line in lines:
        if line.lower().startswith("question:"):
            if current:
                questions.append(_build_word_question(current))
            current = {
                "text": line.split(":", 1)[1].strip(),
                "type": "single_choice",
                "difficulty": "medium",
                "explanation": "",
                "tags": [],
                "options": [],
                "correct_indexes": [],
            }
            continue
        if current is None:
            continue
        lower = line.lower()
        if lower.startswith("type:"):
            current["type"] = line.split(":", 1)[1].strip()
        elif lower.startswith("difficulty:"):
            current["difficulty"] = line.split(":", 1)[1].strip()
        elif lower.startswith("tags:"):
            current["tags"] = [part.strip() for part in line.split(":", 1)[1].split(",") if part.strip()]
        elif lower.startswith("explanation:"):
            current["explanation"] = line.split(":", 1)[1].strip()
        elif lower.startswith("correct:"):
            current["correct_indexes"] = [
                int(part.strip()) - 1
                for part in line.split(":", 1)[1].split(",")
                if part.strip().isdigit()
            ]
        elif lower.startswith(("a.", "b.", "c.", "d.", "e.", "f.", "1.", "2.", "3.", "4.", "5.", "6.")):
            current["options"].append(line.split(".", 1)[1].strip())
    if current:
        questions.append(_build_word_question(current))
    return questions


def _build_word_question(payload: dict) -> BankQuestionCreate:
    options_text = [str(text).strip() for text in payload["options"] if str(text).strip()]
    if len(options_text) < 2:
        raise ValueError(f"Question '{payload['text']}' must include at least 2 options")
    correct_indexes = set(int(index) for index in payload["correct_indexes"])
    if not correct_indexes:
        raise ValueError(f"Question '{payload['text']}' must define Correct: indexes")
    return BankQuestionCreate(
        question_type=_parse_question_type(payload["type"]),
        text=str(payload["text"]).strip(),
        explanation=str(payload["explanation"]).strip() or None,
        difficulty=_parse_difficulty(payload["difficulty"]),
        is_active=True,
        options=[
            BankAnswerOptionCreate(
                order_index=index,
                text=text,
                is_correct=index in correct_indexes,
            )
            for index, text in enumerate(options_text)
        ],
        tags=list(payload["tags"]),
    )


def parse_question_import(filename: str, content: bytes) -> list[BankQuestionCreate]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return parse_excel_questions(content)
    if lower.endswith(".csv"):
        return parse_csv_questions(content)
    if lower.endswith(".docx"):
        return parse_word_questions(content)
    raise ValueError("Only .xlsx, .csv, and .docx files are supported")
